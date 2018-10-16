from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook
import selenium.common.exceptions
from bs4 import BeautifulSoup
from datetime import datetime
import re

_user_agents = ['Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/65.0.3325.181 Safari/537.36']

chrome_options = Options()
chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
chrome_driver = "/usr/local/bin/chromedriver"
driver = webdriver.Chrome(chrome_driver, chrome_options=chrome_options)

print (driver.title)
path = ''
file_name = "amazon_url.xlsx"
wb = load_workbook(path + file_name)
ws = wb['Sheet1']

row = 2
page = 1
stop =5

while True:
    url = ws.cell(row=int(row), column=2).value
    soup = BeautifulSoup(driver.page_source, 'lxml')
    a = datetime.now()
    driver.get(url)


    try:
        used_box = soup.find_all(id="usedbuyBox")
        if len(used_box)>0:
            print("It is used")
            for parent in soup.find_all(id="usedbuyBox"):
                    for shop_used in parent.find_all(class_="a-link-normal", href=re.compile('help/seller')):
                        shop_used_link = shop_used.get('href')
                        useller_id = shop_used_link.split('UTF8seller=')[0]
                        print(useller_id)
                        ws.cell(row=int(row),column=3).value = shop_used.text
                        ws.cell(row=int(row),column=4).value = 'https://www.amazon.co.jp' + shop_used_link
                        ws.cell(row=int(row),column=5).value = 'Done!'
                        row+=1
                        print(shop_used.text)
                        print(shop_used)
                        print('done')
                        wb.save(path+file_name)
                        print(page)
                        page +=1

        else:
            print("Its New item")
            for ssouce in soup.find_all(id="merchant-info"):
                for seller in ssouce.find_all('a'):
                    sellers = seller.get('href')
                    print(sellers)
                    sellers_id = sellers.split('UTF8seller=')[0]
                    print(sellers_id)
                    ws.cell(row=int(row),column=3).value = seller.text
                    ws.cell(row=int(row),column=4).value = 'https://www.amazon.co.jp' + sellers
                    ws.cell(row=int(row),column=5).value = 'Done!'
                    row+=1
                    print('Done')
                    wb.save(path+file_name)
                    print(page)
                    page +=1

    except selenium.common.exceptions.NoSuchElementException:
        pass
        ws.cell(row=int(row),column=3).value = "Error"
        print('Error')
        row+=1

    if  page == stop :
        break



