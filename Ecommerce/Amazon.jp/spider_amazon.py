from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook
import selenium.common.exceptions
from bs4 import BeautifulSoup
from datetime import datetime
import re

_user_agents = ['Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/65.0.3325.181 Safari/537.36']

chrome_options = Options()
chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
chrome_driver = "/usr/local/bin/chromedriver"
driver = webdriver.Chrome(chrome_driver, chrome_options=chrome_options)

print (driver.title)
path = ''
file_name = "amazon_spider.xlsx"
wb = load_workbook(path + file_name)
ws = wb['Sheet1']

row = 2
page = 1
stop = 4

while True:
    url = ws.cell(row=int(row), column=2).value
    soup = BeautifulSoup(driver.page_source, 'lxml')
    a = datetime.now()

    driver.get(url)

    try:
        for shop_info in soup.find_all(class_="a-unordered-list a-nostyle a-vertical"):
            com1 = shop_info.find('span', class_="a-text-bold", text=re.compile('販売業者:'))
            name1 = shop_info.find('span', class_="a-text-bold", text=re.compile('お問い合わせ先電話番号:'))
            add1 = shop_info.find('span', class_="a-text-bold", text=re.compile('住所:'))
            pic1 = shop_info.find('span', class_="a-text-bold", text=re.compile('運営責任者名:'))
            sn1 = shop_info.find('span', class_="a-text-bold", text=re.compile('店舗名:'))
            lic1 = shop_info.find('span', class_="a-text-bold", text=re.compile('古物商許可証番号:'))

            company = com1.next_sibling
            phone = name1.next_sibling
            street = add1.next_sibling.get_text()
            pic = pic1.next_sibling
            shop_name = sn1.next_sibling
            lic = lic1.next_sibling

            print(company)
            print(phone)
            print(street)
            print(pic)
            print(shop_name)
            print(lic)

            ws.cell(row=row, column=3).value = shop_name
            ws.cell(row=row, column=4).value = company
            ws.cell(row=row, column=5).value = street
            ws.cell(row=row, column=6).value = phone
            ws.cell(row=row, column=7).value = pic
            ws.cell(row=row, column=8).value = lic

            row +=1
            page +=1
            wb.save(file_name)


    except AttributeError:
        pass
        wb.save(file_name)
        print('Error')
        row +=1
        page +=1


    if  page == stop :
        break



