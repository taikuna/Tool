from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook
import selenium.common.exceptions
from bs4 import BeautifulSoup
from datetime import datetime
import re
from selenium.webdriver.common.action_chains import ActionChains

_user_agents = ['Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/65.0.3325.181 Safari/537.36']

chrome_options = Options()
chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
chrome_driver = "/usr/local/bin/chromedriver"
driver = webdriver.Chrome(chrome_driver, options=chrome_options)

timeout = 5
print (driver.title)
path = ''
file_name = "client_finder_amazon.xlsx"
wb = load_workbook(path + file_name)
ws = wb['Sheet1']

url = 'https://www.amazon.co.jp/s/ref=sr_pg_2?rh=n%3A2127209051%2Ck%3Amacbook&page=1'

driver.get(url)

a = datetime.now()
soup = BeautifulSoup(driver.page_source, 'lxml')

row = 2
rowc = 2
earow = 2

for res_list in soup.find_all(id="s-results-list-atf"):
    for link in res_list.find_all(class_="a-link-normal s-access-detail-page s-color-twister-title-link a-text-normal"):
        links = link.get('href')
        timenow = a.ctime()
        ws.cell(row = row, column = 2).value = links
        ws.cell(row = row, column = 1).value = timenow
        row+= 1
        wb.save(file_name)



while True:

    eapage = ws.cell(row=int(earow), column=2).value
    driver.get(eapage)

    print (driver.title)

    sh_pg = driver.find_element_by_xpath('//*[@id="merchant-info"]/a')
    ActionChains(driver).move_to_element(sh_pg).click().perform()

    driver.implicitly_wait(5)

    soup = BeautifulSoup(driver.page_source, 'lxml')

    for shop_info in soup.find_all(class_="a-unordered-list a-nostyle a-vertical"):
        com1 = shop_info.find(class_='a-list-item').next_sibling
        name1 = shop_info.find(class_="a-text-bold", text=re.compile('お問い合わせ先電話番号:'))
        add1 = shop_info.find(lass_="a-text-bold", text=re.compile('住所:'))
        pic1 = shop_info.find(class_="a-text-bold", text=re.compile('運営責任者名:'))
        sn1 = shop_info.find(class_="a-text-bold", text=re.compile('店舗名:'))
        lic1 = shop_info.find(class_="a-text-bold", text=re.compile('古物商許可証番号:'))

        company = com1
        ws.cell(row=earow, column=4).value = company
        print(company)

        if name1 is not None:
            phone = name1.next_sibling
            ws.cell(row=earow, column=6).value = phone
            print(phone)
        else:
            ws.cell(row=earow, column=6).value = ''

        if sn1 is not None:
            shop_name = sn1.next_sibling
            ws.cell(row=earow, column=3).value = shop_name
            print(shop_name)
        else:
            ws.cell(row=earow, column=3).value = ''

        if add1 is not None:
            street = add1.next_sibling.get_text()
            ws.cell(row=earow, column=5).value = street
            print(street)
        else:
            ws.cell(row=earow, column=5).value = ''

        if lic1 is not None:
            lic = lic1.next_sibling
            ws.cell(row=earow, column=8).value = lic
            print(lic)
        else:
            ws.cell(row=earow, column=8).value = ''

        if pic1 is not None:
            pic = pic1.next_sibling
            ws.cell(row=earow, column=7).value = pic
            print(pic)
        else:
            ws.cell(row=earow, column=7).value = ''

        earow +=1
        wb.save(file_name)
        print(earow)
