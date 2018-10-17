from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook
import selenium.common.exceptions
from bs4 import BeautifulSoup
from datetime import datetime
import re
from selenium.webdriver.common.action_chains import ActionChains

_user_agents = ['Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/65.0.3325.181 Safari/537.36']

chrome_options = Options()
chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
chrome_driver = "/usr/local/bin/chromedriver"
driver = webdriver.Chrome(chrome_driver, options=chrome_options)

print (driver.title)
path = ''
file_name = "amazon_recorder.xlsx"
wb = load_workbook(path + file_name)
ws = wb['Sheet1']

rurl = 'https://www.amazon.co.jp/s?marketplaceID=&merchant='
seller_id = 'A2NI8ZJ1TR3GGX'

url = (rurl+seller_id)

driver.get(url)

a = datetime.now()

page = 1
stop =10

row = 2
while True:
    timenow = a.ctime()
    soup = BeautifulSoup(driver.page_source, 'lxml')
    for list_atf in soup.find_all(id="s-results-list-atf"):
        for sp_mini in list_atf.find_all(class_="a-row a-spacing-mini"):
            for s_inline in sp_mini.find_all(class_="a-size-base s-inline s-access-title a-text-normal"):
                title = s_inline.get_text()
                print(title)
            for link in sp_mini.find_all('a', class_="a-link-normal s-access-detail-page s-color-twister-title-link a-text-normal"):
                links = link.get('href')
                print(links)
            for aprice in sp_mini.find_all(class_="a-size-base a-color-price s-price a-text-bold"):
                price = aprice.get_text()
                print(price)
            for astock in sp_mini.find_all(class_="a-size-small a-color-price"):
                stock = astock.get_text()
                print(stock)

                ws.cell(row=int(row), column=1).value = timenow
                ws.cell(row=int(row), column=2).value = title
                ws.cell(row=int(row), column=5).value = links
                ws.cell(row=int(row), column=3).value = price
                ws.cell(row=int(row), column=4).value = stock
                row +=1
                page +=1
                wb.save(file_name)

                next = driver.find_element_by_xpath('//*[@class="srSprite pagnNextArrow"]')
                try :
                    ActionChains(driver).move_to_element(next).click().perform()
                except selenium.common.exceptions.StaleElementReferenceException:
                    pass
                    next.click
                    wb.save(file_name)



    if page == stop:
        break