# -*- coding: utf-8 -*-
from bs4 import BeautifulSoup
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import openpyxl
from selenium.webdriver.common.action_chains import ActionChains
import re

_user_agents = ['Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/65.0.3325.181 Safari/537.36']

chrome_options = Options()
chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
chrome_driver = "/usr/local/bin/chromedriver"
driver = webdriver.Chrome(chrome_driver, chrome_options=chrome_options)

i=1

row = 2
row1 =2

path = ''

file_name = "yahoo.xlsx"
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'test_sheet_1'

ws.cell(row=1, column=1).value = 'Time'
ws.cell(row=1, column=2).value = 'Page title'
ws.cell(row=1, column=3).value = 'Hit'
ws.cell(row=1, column=4).value = 'Item'
ws.cell(row=1, column=5).value = 'URL'
ws.cell(row=1, column=6).value = 'Price'
ws.cell(row=1, column=7).value = 'Seller'


page = 1
row = 2
row1 = 2
row2 = 2
row3 =2

url= 'https://shopping.yahoo.co.jp/category/2498/list?oq=&uIv=on&pf=&pt=&used=0&seller=0&ei=UTF-8&b=' + str(page)
driver.get(url)

while True:
    soup = BeautifulSoup(driver.page_source, 'lxml')
    a = datetime.now()
    title = soup.head.title
    for pseller in soup.find_all(class_="elSeller"):
        for seller in pseller.find_next_siblings('span'):
                ws.cell(row=row2, column=7).value = seller.text
                row2 +=1

    for parent in soup.find_all(class_='mdSearchList elList'):
        for item in parent.find_all('dd', class_="elName", string=True):
            for link in item.find_all('a'):
                ws.cell(row=row, column=4).value = item.string
                ws.cell(row=row, column=5).value = link.get('href')
                ws.cell(row=row, column=1).value = a.ctime()
                ws.cell(row=row, column=2).value = title.text
                ws.cell(row=row, column=3).value = 'page'+str(page)
                row +=1

    for pricea in soup.find_all(class_='mdSearchList elList'):
        for priceb in pricea.find_all(class_="elUnit"):
            for price in priceb.find_previous_siblings('span'):
                ws.cell(row=row1, column=6).value = price.text

                row1 +=1

    page +=1
    print(title.text)
    wb.save(path + file_name)
    next = driver.find_element_by_xpath('//*[@class="elNext"]')
    ActionChains(driver).move_to_element(next).click().perform()
    print('page'+ str(page))

    stop = 15

    if page == stop:
        break














