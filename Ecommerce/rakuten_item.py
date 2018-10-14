# -*- coding: utf-8 -*-
from bs4 import BeautifulSoup
import requests
from openpyxl import load_workbook
import cchardet
from datetime import datetime

row=2
path = ''
file_name = "rakuten_item.xlsx"
wb = load_workbook(path + file_name)
ws = wb['Sheet1']
p = 1
ws.cell(row=1, column=1).value ='Time'
ws.cell(row=1, column=2).value ='Page Title'
ws.cell(row=1, column=3).value ='Page count'
ws.cell(row=1, column=4).value ='Item Title'
ws.cell(row=1, column=5).value ='Item URL'
ws.cell(row=1, column=6).value ='Price'

url = "https://search.rakuten.co.jp/search/mall/+%E5%A5%B3%E6%80%A7%E7%94%A8%E3%80%80%E3%82%B9%E3%82%AB%E3%83%BC%E3%83%95/?p="+str(p)

while True:
    try:
        print(p)

        r = requests.get(url)
        r.encoding = cchardet.detect(r.content)["encoding"]
        soup = BeautifulSoup(r.text, 'lxml')

        keyword_hit = soup.find(class_="_big section keyword _break").text
        a = datetime.now()
        print(keyword_hit)

        sp_count = soup.find(class_="count _medium").text
        span_count = sp_count.replace('\n', '').replace(' ', '')

        for important in soup.find_all('span', class_="important"):
            ws.cell(row=row, column=6).value = important.text
            print(important.text)


        for item_list in soup.find_all(class_="content title"):
             for item_link in item_list.find_all('a', href=True):
                ws.cell(row=row, column=5).value = item_link.get('href')
                ws.cell(row=row, column=4).value = item_link.text
                ws.cell(row=row, column=3).value = span_count
                ws.cell(row=row, column=2).value = keyword_hit
                ws.cell(row=row, column=1).value = a.ctime()
                row += 1
                p +=1
                wb.save(path + file_name)

    except(KeyError):
        pass
        print("error")


    wb.save(path + file_name)
    if p ==3:
        break

