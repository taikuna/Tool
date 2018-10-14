# -*- coding: utf-8 -*-
from bs4 import BeautifulSoup
import requests
from openpyxl import load_workbook
import cchardet
from datetime import datetime

rl=2
rt=2
i=5
path = ''
file_name = "rakuten.xlsx"
wb = load_workbook(path + file_name)
ws = wb['Sheet1']

ws.cell(row=1, column=1).value = 'Time'

while True:
    print(i)
    url = "https://search.rakuten.co.jp/search/mall/-/303656/?p=" + str(i)
    r = requests.get(url)
    r.encoding = cchardet.detect(r.content)["encoding"]
    soup = BeautifulSoup(r.text, 'lxml')
    pg = soup.head.title

    thread = soup.find(class_="count _medium").text
    th_sp = thread.replace('\n', '').replace(' ', '')
    a = datetime.now()
    print(th_sp)


    for shop_list in soup.find_all(class_="content merchant _ellipsis"):
        for thread_link in shop_list.find_all('a', href=True):
            ws.cell(row=rl, column=5).value = thread_link['href']
            ws.cell(row=rl, column=2).value = pg.text
            print(thread_link.get('href'))
            rl += 1

    for shop_title in soup.find_all(class_="content merchant _ellipsis"):
        for thread_title in shop_title.find_all('a', href=True):
            ws.cell(row=rt, column=4).value = thread_title.text
            ws.cell(row=rt, column=3).value = th_sp
            ws.cell(row=rt, column=1).value = a.ctime()
            print(thread_title.text)
            rt += 1

    wb.save(path + file_name)
    i+=1
    print(pg)
    print(rl)
    print(rt)
    print(i)
    if rl > 5:
        break

