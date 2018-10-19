from bs4 import BeautifulSoup
import requests
from openpyxl import load_workbook
import cchardet
from datetime import datetime
import re
from time import sleep
from requests import get

path = ''
file_name = "town_page.xlsx"
wb = load_workbook(path + file_name)
ws = wb['Sheet2']
a = datetime.now()

link = 'https://itp.ne.jp/result/?kw=%E4%B8%8D%E5%8B%95%E7%94%A3&sr=1&evdc=1&num=50&pg='
def orange_text():
    page = 3
    row = 2
    stop = 6

    while True:
        print("Page "+ str(page)+' Scan start....')
        url = link + str(page)
        r = requests.get(url)
        r.encoding = cchardet.detect(r.content)["encoding"]
        soup = BeautifulSoup(r.text, 'lxml')
        timenow = a.ctime()
        for title in soup.find_all(class_="searchResultHeader"):
            print(title.text)
            for result in soup.find_all(class_="normalResultsBox"):
                for name2 in result.find_all(class_="orangeText"):

                    name_black = name2.get_text()
                    id = name2.get('href').split('/')[2]
                    print(name_black)
                    print(id)
                    ws.cell(row = row, column = 1).value = timenow
                    ws.cell(row = row, column = 4).value = name_black
                    ws.cell(row = row, column = 6).value = id
                    ws.cell(row = row, column = 2).value = title.text
                    ws.cell(row = row, column = 3).value = 'page'+str(page)
                    row +=1
                    wb.save(file_name)
                    print('page '+str(page)+' scan done, workbook saved')
                    print('row ' + str(row))

        sleep(2)
        page +=1

        if str(page) >= str(stop):
            break

def blue_text():

    page= ws.cell(row = 1, column = 2).value
    row = ws.cell(row = 1, column = 1).value
    stop = 0

    while True:
        print("Search Page"+ str(page))
        url = link + str(page)
        r = requests.get(url)
        r.encoding = cchardet.detect(r.content)["encoding"]
        soup = BeautifulSoup(r.text, 'lxml')
        timenow = a.ctime()
        for title in soup.find_all(class_="searchResultHeader"):
            print(title.text)
            print('Its Blue')
            for result in soup.find_all(class_="normalResultsBox"):
                for name2 in result.find_all(class_="blueText"):
                    name_black = name2.get_text()
                    id = name2.get('href').split('/')[2]
                    print(name_black)
                    print(id)

                    ws.cell(row = row, column = 1).value = timenow
                    ws.cell(row = row, column = 4).value = name_black
                    ws.cell(row = row, column = 5).value = id
                    ws.cell(row = row, column = 2).value = title.text
                    ws.cell(row = row, column = 3).value = 'page'+str(page)
                    row +=1
                    wb.save(file_name)
        sleep(3)
        page +=1
        print('page'+str(page))
        if page == stop:
            break

def black_text():
    page=38
    row =2
    stop = 0

    while True:
        print("Search Page"+ str(page))
        url = link + str(page)
        r = requests.get(url)
        r.encoding = cchardet.detect(r.content)["encoding"]
        soup = BeautifulSoup(r.text, 'lxml')
        timenow = a.ctime()
        for title in soup.find_all(class_="searchResultHeader"):
            print(title.text)
            print('Its Black')
            for result in soup.find_all(class_="normalResultsBox"):
                for name2 in result.find_all(class_="blackText"):
                    name_black = name2.get_text()
                    id = name2.get('href').split('/')[2]
                    print(name_black)
                    print(id)

                for hp2 in result.find_all(class_="inlineSmallHeader",text=re.compile('URL')):
                    hp_black = hp2.next_sibling.replace(' ','').replace('\n','')
                    print(hp_black)
                    ws.cell(row = row, column = 1).value = timenow
                    ws.cell(row = row, column = 4).value = name_black
                    ws.cell(row = row, column = 6).value = hp_black
                    ws.cell(row = row, column = 5).value = id
                    ws.cell(row = row, column = 2).value = title.text
                    ws.cell(row = row, column = 3).value = 'page'+str(page)
                    row +=1
                    wb.save(file_name)
        sleep(3)
        page +=1
        print('page'+str(page))
        if page == stop:
            break

def finder():
    row = 2

    while True:
        print(row)
        info_id = ws.cell(row=int(row), column=5).value
        url = 'https://itp.ne.jp/info/' + info_id + '/shop/'

        response = get(url)
        response.encoding = response.apparent_encoding
        data = response.text
        soup = BeautifulSoup(data, 'lxml')
        emails = set(re.findall(r"[a-z0-9\.\-+_]+@[a-z0-9\.\-+_]+\.[a-z]+", response.text, re.I))
        emails = set()
        website = set(soup.find_all('a', text=re.compile('http')))
        website =set()

        columnw = 10
        columne =7
        for emails in set(re.findall(r"[a-z0-9\.\-+_]+@[a-z0-9\.\-+_]+\.[a-z]+", response.text, re.I)):
            ws.cell(row=row, column=columne).value = emails
        for website in set(soup.find_all('a', text=re.compile('http'))):
            ws.cell(row=row, column=columnw).value = website.text
            columnw += 1
            print(website.text)

        row +=1

        print(soup.title.string)
        print(emails)
        wb.save(file_name)


        if row == 0:
            break

orange_text()
