# -*- coding: utf-8 -*-
from bs4 import BeautifulSoup
import re
from requests import get
from openpyxl import load_workbook

path = ''
file_name = "contact_list.xlsx"
wb = load_workbook(path + file_name)
ws = wb['Contact']

row = 2

while True:
    print(row)
    url = ws.cell(row=int(row), column=2).value
    response = get(url)
    response.encoding = response.apparent_encoding
    data = response.text
    soup = BeautifulSoup(data, 'lxml')
    str_cmp_list = ['お問い合わせ', '問合せ', '連絡', '問い合せ', 'contact']
    str_cmp = re.compile('|'.join(str_cmp_list))
    tags = soup.find_all("a", string=str_cmp)
    links = soup.find_all("a", href=str_cmp)
    emails = set(re.findall(r"[a-z0-9\.\-+_]+@[a-z0-9\.\-+_]+\.[a-z]+", response.text, re.I))
    emails = set()
    column = 5

    for cus in soup.find_all(r"a", string=str_cmp):
        ws.cell(row=row, column=column).value = cus['href']
        column += 1

    for link in soup.find_all(r"a", href=str_cmp):
        ws.cell(row=row, column=column).value = link['href']
        column += 1

    for emails in set(re.findall(r"[a-z0-9\.\-+_]+@[a-z0-9\.\-+_]+\.[a-z]+", response.text, re.I)):
        ws.cell(row=row, column=10).value = emails
        column += 1

    ws.cell(row=row, column=3).value = soup.title.string
    ws.cell(row=row, column=4).value = url
    row +=1
    print(soup.title.string)
    print(url)
    print(tags)
    print(links)
    print(emails)
    wb.save(path + file_name)
    if row > 400:
        break